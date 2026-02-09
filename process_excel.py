import pandas as pd
import os
import shutil
import argparse
from datetime import datetime

# 命令行参数解析
def parse_args():
    parser = argparse.ArgumentParser(description='处理Excel文件并计算更新数据')
    parser.add_argument('source_file', help='源文件路径')
    parser.add_argument('target_file', help='目标文件路径')
    return parser.parse_args()

# 获取带时间戳的文件名
def get_timestamped_filename(file_path):
    directory = os.path.dirname(file_path)
    filename = os.path.basename(file_path)
    name, ext = os.path.splitext(filename)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return os.path.join(directory, f'{name}_{timestamp}{ext}')

# 主函数
def extract_model(code):
    """从编码中提取产品型号"""
    if isinstance(code, str) and '-' in code:
        parts = code.split('-')
        if len(parts) >= 2:
            return '-'.join(parts[1:])
    return None

def main():
    # 解析命令行参数
    args = parse_args()
    source_file = args.source_file
    target_file = args.target_file
    
    print(f"源文件: {source_file}")
    print(f"目标文件: {target_file}")
    
    # 验证文件存在
    if not os.path.exists(source_file):
        print(f"错误：源文件不存在: {source_file}")
        return
    if not os.path.exists(target_file):
        print(f"错误：目标文件不存在: {target_file}")
        return
    
    # 读取源文件
    print(f"读取源文件: {source_file}")
    try:
        # 注意：虽然文件扩展名是.csv，但实际是Excel格式
        # 使用第二行作为列名（索引为1）
        df_source = pd.read_excel(source_file, header=1)
        print(f"成功读取源文件，共 {len(df_source)} 行数据")
        print("源文件列名:")
        for i, col in enumerate(df_source.columns):
            print(f"{i}: {col}")
            # 打印每列的前几个值
            print(f"  前5个值: {df_source[col].head().tolist()}")
    except Exception as e:
        print(f"读取源文件失败: {e}")
        return
    
    # 提取产品型号
    print("提取产品型号...")
    # 尝试常见的列名
    merchant_code_cols = [col for col in df_source.columns if '商家' in col and '编码' in col]
    print(f"找到的商家编码列: {merchant_code_cols}")
    
    for col in merchant_code_cols:
        df_source[f'产品型号_{col}'] = df_source[col].apply(extract_model)
    
    # 确定最终产品型号
    model_cols = [col for col in df_source.columns if '产品型号_' in col]
    print(f"找到的产品型号列: {model_cols}")
    
    if model_cols:
        # 使用第一个找到的产品型号列
        df_source['产品型号'] = df_source[model_cols[0]]
        # 如果有多个型号列，使用fillna合并
        for col in model_cols[1:]:
            df_source['产品型号'] = df_source['产品型号'].fillna(df_source[col])
    else:
        print("警告：未能提取任何产品型号")
        return
    

    
    # 计算差值
    print("计算差值...")
    if '实际可用数' in df_source.columns and '30天销量' in df_source.columns:
        df_source['差值'] = df_source['实际可用数'] - df_source['30天销量']
    else:
        print("警告：未找到'实际可用数'或'30天销量'列")
        return
    
    # 生成带时间戳的输出文件名
    output_file = get_timestamped_filename(target_file)
    print(f"复制原始文件到: {output_file}")
    try:
        shutil.copy2(target_file, output_file)
        print("文件复制成功，保留了原始格式和图片")
    except Exception as e:
        print(f"复制文件失败: {e}")
        return
    
    # 使用openpyxl的不同方法打开文件，尝试保留图片
    print(f"使用openpyxl打开文件: {output_file}")
    try:
        # 使用openpyxl打开文件，启用所有选项
        from openpyxl import load_workbook
        wb = load_workbook(output_file, data_only=False, keep_links=True)
        ws = wb.active
        print(f"成功打开文件，工作表名称: {ws.title}")
        print(f"文件包含 {ws.max_row} 行, {ws.max_column} 列")
        print(f"文件包含 {len(ws._images)} 个图片")
    except Exception as e:
        print(f"打开文件失败: {e}")
        return
    
    # 智能识别产品型号列和所需数量列
    print("智能识别列...")
    
    # 查找产品型号列（包含'产品型号'值的列）
    product_model_col_idx = None
    required_qty_col_idx = None
    
    # 遍历前5行查找列标题
    for row_idx in range(1, 6):  # 遍历前5行
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if isinstance(cell_value, str):
                if '产品型号' in cell_value and not product_model_col_idx:
                    product_model_col_idx = col_idx
                    print(f"在第 {row_idx} 行找到产品型号列，列索引: {product_model_col_idx}")
                elif '所需数量' in cell_value and not required_qty_col_idx:
                    required_qty_col_idx = col_idx
                    print(f"在第 {row_idx} 行找到所需数量列，列索引: {required_qty_col_idx}")
        # 如果都找到了，就退出循环
        if product_model_col_idx and required_qty_col_idx:
            break
    
    print(f"识别到的产品型号列索引: {product_model_col_idx}")
    print(f"识别到的所需数量列索引: {required_qty_col_idx}")
    
    if product_model_col_idx and required_qty_col_idx:
        # 合并数据
        print("根据产品型号合并数据...")
        
        # 创建产品型号到差值的映射
        model_diff_map = df_source.set_index('产品型号')['差值'].to_dict()
        print(f"源文件中找到 {len(model_diff_map)} 个产品型号与差值映射")
        
        # 更新目标文件中的所需数量列（从第四行开始，因为第一行是标题，第二行是列名，第三行是分类）
        updated_count = 0
        max_row = ws.max_row
        print(f"开始更新数据，从第4行到第{max_row}行")
        
        # 遍历数据行，只修改所需数量列
        for row in range(4, max_row + 1):
            # 获取产品型号
            model = ws.cell(row=row, column=product_model_col_idx).value
            
            # 只在找到匹配的产品型号时更新
            if model and model in model_diff_map:
                diff_value = model_diff_map[model]
                # 直接写入值，保留原始格式
                ws.cell(row=row, column=required_qty_col_idx).value = diff_value
                print(f"更新产品型号 {model} 的所需数量为 {diff_value}")
                updated_count += 1
        
        print(f"数据更新完成，共更新了 {updated_count} 个单元格")
        print(f"文件中图片数量: {len(ws._images)}")
        
        # 保存更新后的文件
        print(f"保存更新后的文件: {output_file}")
        wb.save(output_file)
        print(f"文件更新成功！共更新了 {updated_count} 个产品型号")
        print("提示：文件是通过复制原始文件后修改的，图片数据应该已经保留")
        print("注意：openpyxl可能无法正确显示嵌入图片，但图片数据应该仍然存在于文件中")
    else:
        print("错误：未找到合适的产品型号列或所需数量列")

if __name__ == "__main__":
    main()