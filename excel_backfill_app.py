import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import os
from difflib import SequenceMatcher
import xlrd
from copy import copy
import tempfile

st.set_page_config(page_title="Excel数据回填工具", layout="wide")

st.title("Excel数据回填工具")
st.markdown("将源Excel数据回填到目标模板，保留模板格式和样式")

def detect_header_row(df):
    for i in range(min(10, len(df))):
        row = df.iloc[i]
        non_null_count = row.notna().sum()
        if non_null_count >= len(row) * 0.7:
            return i
    return 0

def detect_data_start_row(df, header_row):
    for i in range(header_row + 1, min(header_row + 10, len(df))):
        row = df.iloc[i]
        if row.notna().sum() >= 2:
            return i
    return header_row + 1

def similarity(a, b):
    if pd.isna(a) or pd.isna(b):
        return 0
    return SequenceMatcher(None, str(a).lower(), str(b).lower()).ratio()

def auto_match_columns(source_cols, target_cols):
    mapping = {}
    
    keyword_mappings = {
        '商品编码': ['型号', '编码', '货号', 'sku', 'code'],
        '采购数量': ['数量', '采购数', 'qty', 'quantity'],
        '单价': ['单价', '价格', 'price'],
        '采购金额': ['总价', '金额', 'amount', 'total'],
        '供应商': ['店铺', '供应商', 'vendor', 'supplier'],
        '备注': ['备注', '说明', 'note', 'remark'],
        '品名': ['品名', '名称', 'name', 'product'],
    }
    
    for target_col in target_cols:
        if pd.isna(target_col):
            continue
        target_str = str(target_col).strip()
        best_match = None
        best_score = 0
        
        if target_str in keyword_mappings:
            keywords = keyword_mappings[target_str]
            for source_col in source_cols:
                if pd.isna(source_col):
                    continue
                source_str = str(source_col).strip().lower()
                for keyword in keywords:
                    if keyword.lower() in source_str:
                        score = 0.9
                        if score > best_score and source_col not in mapping.values():
                            best_score = score
                            best_match = source_col
        
        if best_match is None:
            for source_col in source_cols:
                if pd.isna(source_col):
                    continue
                score = similarity(target_str, source_col)
                if score > best_score and score > 0.5 and source_col not in mapping.values():
                    best_score = score
                    best_match = source_col
        
        if best_match:
            mapping[target_str] = best_match
    
    return mapping

def xls_to_xlsx_from_bytes(file_bytes):
    with tempfile.NamedTemporaryFile(suffix='.xls', delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    
    try:
        xls_book = xlrd.open_workbook(tmp_path, formatting_info=True)
        xls_sheet = xls_book.sheet_by_index(0)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                cell = ws.cell(row=row_idx + 1, column=col_idx + 1)
                try:
                    cell.value = xls_sheet.cell_value(row_idx, col_idx)
                except:
                    cell.value = None
                
                try:
                    xf_idx = xls_sheet.cell_xf_index(row_idx, col_idx)
                    xf = xls_book.xf_list[xf_idx]
                    
                    font = Font()
                    try:
                        font_idx = xf.font_index
                        font_data = xls_book.font_list[font_idx]
                        font = Font(
                            bold=font_data.bold != 0,
                            italic=font_data.italic != 0,
                            size=font_data.height / 20 if font_data.height else 11
                        )
                    except:
                        pass
                    cell.font = font
                    
                    try:
                        align = Alignment(
                            horizontal=['left', 'center', 'right'][xf.alignment.hor_align] if xf.alignment.hor_align < 3 else 'left',
                            vertical=['top', 'center', 'bottom'][xf.alignment.vert_align] if xf.alignment.vert_align < 3 else 'center'
                        )
                        cell.alignment = align
                    except:
                        pass
                    
                except:
                    pass
        
        for col_idx in range(xls_sheet.ncols):
            try:
                col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                ws.column_dimensions[col_letter].width = 15
            except:
                pass
        
        return wb
    finally:
        os.unlink(tmp_path)

def load_excel_from_uploaded(uploaded_file):
    file_bytes = uploaded_file.read()
    file_name = uploaded_file.name
    
    if file_name.endswith('.xls'):
        df = pd.read_excel(BytesIO(file_bytes), header=None, engine='xlrd')
        wb = xls_to_xlsx_from_bytes(file_bytes)
    else:
        df = pd.read_excel(BytesIO(file_bytes), header=None)
        wb = openpyxl.load_workbook(BytesIO(file_bytes))
    
    return df, wb

col1, col2 = st.columns(2)

with col1:
    st.subheader("源文件设置")
    source_file = st.file_uploader(
        "上传源文件",
        type=['xlsx', 'xls'],
        key="source_uploader"
    )
    
    if source_file is not None:
        try:
            source_df, source_wb = load_excel_from_uploaded(source_file)
            st.session_state['source_df'] = source_df
            st.session_state['source_wb'] = source_wb
            
            auto_header = detect_header_row(source_df)
            st.session_state['auto_header_row'] = auto_header
            
            auto_data_start = detect_data_start_row(source_df, auto_header)
            st.session_state['auto_data_start'] = auto_data_start
            
            st.success(f"加载成功！共 {len(source_df)} 行，{len(source_df.columns)} 列")
        except Exception as e:
            st.error(f"加载失败: {e}")

with col2:
    st.subheader("目标模板设置")
    target_file = st.file_uploader(
        "上传目标模板",
        type=['xlsx', 'xls'],
        key="target_uploader"
    )
    
    if target_file is not None:
        try:
            target_df, target_wb = load_excel_from_uploaded(target_file)
            st.session_state['target_df'] = target_df
            st.session_state['target_wb'] = target_wb
            st.success(f"加载成功！共 {len(target_df.columns)} 列")
        except Exception as e:
            st.error(f"加载失败: {e}")

if 'source_df' in st.session_state and 'target_df' in st.session_state:
    st.divider()
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("源文件配置")
        
        source_df = st.session_state['source_df']
        
        st.markdown("**预览源文件前10行：**")
        st.dataframe(source_df.head(10), use_container_width=True)
        
        header_row = st.number_input(
            "表头行（从0开始）",
            min_value=0,
            max_value=len(source_df) - 1,
            value=st.session_state.get('auto_header_row', 0),
            key="source_header_row"
        )
        
        data_start_row = st.number_input(
            "数据起始行（从0开始）",
            min_value=0,
            max_value=len(source_df) - 1,
            value=st.session_state.get('auto_data_start', 1),
            key="source_data_start"
        )
        
        source_headers = source_df.iloc[header_row].tolist()
        st.markdown(f"**识别到的表头：** {source_headers}")
        
        source_data = source_df.iloc[data_start_row:].copy()
        source_data.columns = source_headers
        source_data = source_data.reset_index(drop=True)
        
        st.markdown(f"**数据预览（共{len(source_data)}行）：**")
        st.dataframe(source_data.head(5), use_container_width=True)
    
    with col2:
        st.subheader("目标模板配置")
        
        target_df = st.session_state['target_df']
        
        target_header_row = st.number_input(
            "目标表头行（从0开始）",
            min_value=0,
            max_value=max(1, len(target_df) - 1) if len(target_df) > 0 else 0,
            value=0,
            key="target_header_row"
        )
        
        target_data_start = st.number_input(
            "目标数据起始行（从0开始）",
            min_value=0,
            max_value=max(1, len(target_df) - 1) if len(target_df) > 0 else 1,
            value=1,
            key="target_data_start"
        )
        
        if len(target_df) > 0:
            target_headers = target_df.iloc[target_header_row].tolist()
        else:
            target_wb = st.session_state['target_wb']
            ws = target_wb.active
            target_headers = [cell.value for cell in ws[1]]
            target_data_start = 1
        
        st.markdown(f"**目标模板列：** {target_headers}")
    
    st.divider()
    st.subheader("列映射配置")
    
    prefix = st.text_input("商品编码前缀（可选，将添加到商品编码前）", value="", key="code_prefix")
    
    auto_mapping = auto_match_columns(source_headers, target_headers)
    
    col1, col2, col3 = st.columns([2, 1, 2])
    
    with col1:
        st.markdown("**目标列**")
    with col2:
        st.markdown("**映射**")
    with col3:
        st.markdown("**源列**")
    
    mapping_result = {}
    for target_col in target_headers:
        if pd.isna(target_col):
            continue
        target_str = str(target_col).strip()
        
        col1, col2, col3 = st.columns([2, 1, 2])
        
        with col1:
            st.markdown(f"**{target_str}**")
        
        with col2:
            st.markdown("←")
        
        with col3:
            default_value = auto_mapping.get(target_str, "")
            source_options = ["(不映射)"] + [str(c) if pd.notna(c) else f"列{i}" for i, c in enumerate(source_headers)]
            default_index = 0
            if default_value:
                for i, opt in enumerate(source_options):
                    if opt == str(default_value):
                        default_index = i
                        break
            
            selected = st.selectbox(
                f"映射_{target_str}",
                options=source_options,
                index=default_index,
                key=f"map_{target_str}",
                label_visibility="collapsed"
            )
            
            if selected != "(不映射)":
                mapping_result[target_str] = selected
    
    st.divider()
    
    if st.button("执行数据导入", type="primary"):
        try:
            target_wb = st.session_state['target_wb']
            ws = target_wb.active
            
            source_data_clean = source_data.copy()
            
            for col in source_data_clean.columns:
                source_data_clean[col] = source_data_clean[col].ffill()
            
            imported_count = 0
            for idx, row in source_data_clean.iterrows():
                target_row = target_data_start + idx + 1
                
                for target_col_name, source_col_name in mapping_result.items():
                    target_col_idx = None
                    for col_idx, col_name in enumerate(target_headers, 1):
                        if str(col_name).strip() == target_col_name:
                            target_col_idx = col_idx
                            break
                    
                    if target_col_idx is None:
                        continue
                    
                    source_value = row.get(source_col_name)
                    
                    if pd.notna(source_value):
                        cell = ws.cell(row=target_row, column=target_col_idx)
                        if target_col_name == "商品编码" and prefix:
                            cell.value = prefix + str(source_value)
                        else:
                            cell.value = source_value
                
                imported_count += 1
            
            output_buffer = BytesIO()
            target_wb.save(output_buffer)
            output_buffer.seek(0)
            
            st.session_state['output_buffer'] = output_buffer
            st.session_state['imported_count'] = imported_count
            
            st.success(f"成功导入 {imported_count} 行数据！")
            
        except Exception as e:
            st.error(f"导入失败: {e}")
            import traceback
            st.code(traceback.format_exc())
    
    if 'output_buffer' in st.session_state:
        st.divider()
        st.subheader("导出结果")
        
        output_filename = st.text_input(
            "输出文件名",
            value="采购商品导入模板_已填充.xlsx"
        )
        
        st.download_button(
            label="下载填充后的Excel文件",
            data=st.session_state['output_buffer'],
            file_name=output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
        
        st.info(f"已导入 {st.session_state.get('imported_count', 0)} 行数据，模板原有格式和样式已保留")

st.sidebar.markdown("### 使用说明")
st.sidebar.markdown("""
1. **上传文件**：上传源文件和目标模板
2. **配置源文件**：
   - 确认表头行位置
   - 确认数据起始行位置
3. **配置列映射**：
   - 系统自动匹配列名
   - 可手动调整映射关系
4. **执行导入**：点击执行数据导入
5. **下载结果**：下载填充后的Excel文件
""")

st.sidebar.markdown("### 列映射建议")
st.sidebar.markdown("""
- **商品编码** ← 型号
- **采购数量** ← 数量
- **单价** ← 单价
- **采购金额** ← 总价
- **供应商** ← 店铺名称
""")

st.sidebar.markdown("### 功能说明")
st.sidebar.markdown("""
- 自动向下填充源文件中的空值（如店铺名称）
- 保留目标模板的格式和样式
- 支持 .xls 和 .xlsx 格式
""")
