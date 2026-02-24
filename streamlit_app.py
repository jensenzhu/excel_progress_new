import streamlit as st
import pandas as pd
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, Border, Alignment, Protection
from openpyxl.utils import get_column_letter
import tempfile
import os
import shutil
import difflib

def convert_xls_to_xlsx_with_format(xls_content):
    """将 .xls 文件内容转换为 .xlsx 格式，尽可能保留格式"""
    try:
        import xlrd
        xls_book = xlrd.open_workbook(file_contents=xls_content, formatting_info=True)
        xls_sheet = xls_book.sheet_by_index(0)
    except Exception as e:
        raise Exception(f"无法读取 .xls 文件: {str(e)}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = xls_sheet.name if xls_sheet.name else "Sheet1"
    
    xf_list = xls_book.xf_list
    font_list = xls_book.font_list
    
    for row_idx in range(xls_sheet.nrows):
        xls_row = xls_sheet.row(row_idx)
        for col_idx in range(xls_sheet.ncols):
            cell = xls_sheet.cell(row_idx, col_idx)
            value = cell.value
            
            if value is not None and value != '':
                ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)
                
                try:
                    xf_index = cell.xf_index
                    if xf_index < len(xf_list):
                        xf = xf_list[xf_index]
                        font_index = xf.font_index
                        if font_index < len(font_list):
                            xls_font = font_list[font_index]
                            new_font = Font(
                                bold=xls_font.bold,
                                italic=xls_font.italic,
                                name=xls_font.name,
                                size=xls_font.height / 20 if xls_font.height else 11
                            )
                            ws.cell(row=row_idx + 1, column=col_idx + 1).font = new_font
                except Exception:
                    pass
    
    for col_idx in range(xls_sheet.ncols):
        try:
            col_width = xls_sheet.computed_column_width(col_idx)
            if col_width:
                ws.column_dimensions[get_column_letter(col_idx + 1)].width = col_width / 256.0 * 7
        except Exception:
            pass
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def detect_column_info(ws):
    """智能识别订单表的列信息和数据起始行"""
    product_model_keywords = ['产品型号', '商品货号', '货号', '型号', 'model', 'code']
    target_column_keywords = ['所需数量', '数量', '订货数量', '进货数量', '数量/个', 'quantity', 'qty']
    
    product_model_col_idx = None
    target_col_idx = None
    header_row_idx = None
    data_start_row = None
    
    for row_idx in range(1, min(11, ws.max_row + 1)):
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if isinstance(cell_value, str):
                if not product_model_col_idx:
                    for keyword in product_model_keywords:
                        if keyword in cell_value:
                            product_model_col_idx = col_idx
                            header_row_idx = row_idx
                            break
                
                if not target_col_idx:
                    for keyword in target_column_keywords:
                        if keyword in cell_value:
                            target_col_idx = col_idx
                            if not header_row_idx:
                                header_row_idx = row_idx
                            break
    
    if header_row_idx:
        data_start_row = header_row_idx + 1
        for row_idx in range(header_row_idx + 1, min(header_row_idx + 5, ws.max_row + 1)):
            has_data = False
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None and cell_value != '':
                    has_data = True
                    break
            if has_data:
                data_start_row = row_idx
                break
    
    return {
        'product_model_col_idx': product_model_col_idx,
        'target_col_idx': target_col_idx,
        'header_row_idx': header_row_idx,
        'data_start_row': data_start_row
    }

def get_column_name(ws, col_idx, row_idx):
    """获取指定列在指定行的名称"""
    cell_value = ws.cell(row=row_idx, column=col_idx).value
    return str(cell_value) if cell_value else f'列{col_idx}'

st.set_page_config(
    page_title="Excel数据处理工具",
    page_icon="📊",
    layout="wide"
)

st.title("📊 Excel数据处理工具")
st.markdown("---")

st.markdown("### 📁 文件上传")

col1, col2 = st.columns(2)

with col1:
    st.markdown("#### ERP库存表（from文件）")
    from_file = st.file_uploader(
        "上传ERP库存表",
        type=['xlsx', 'xls', 'csv'],
        key='from_file',
        help="上传包含库存数据的Excel文件"
    )

with col2:
    st.markdown("#### 订单表（dist文件）")
    dist_file = st.file_uploader(
        "上传订单表",
        type=['xlsx', 'xls'],
        key='dist_file',
        help="上传需要更新的订单Excel文件（支持.xlsx和.xls格式）"
    )

st.markdown("---")

st.markdown("### ⚙️ 处理配置")

if dist_file:
    st.markdown("#### 📋 订单表列识别")
    
    dist_file_ext = dist_file.name.lower().split('.')[-1] if dist_file.name else 'xlsx'
    
    if dist_file_ext == 'xls':
        st.warning("⚠️ 检测到 .xls 格式，建议先手动转换为 .xlsx 格式以完整保留样式")
        st.info("💡 转换方法：在 Excel 中打开文件，选择'文件 > 另存为 > Excel 工作簿 (.xlsx)'")
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_preview:
        if dist_file_ext == 'xls':
            st.info("🔄 正在转换 .xls 为 .xlsx...")
            try:
                dist_content = dist_file.getvalue()
                converted_content = convert_xls_to_xlsx_with_format(dist_content)
                tmp_preview.write(converted_content)
                st.success("✅ 转换成功！")
            except Exception as e:
                st.error(f"❌ 转换失败: {str(e)}")
                os.unlink(tmp_preview.name)
                st.stop()
        else:
            tmp_preview.write(dist_file.getvalue())
        tmp_preview_path = tmp_preview.name
    
    try:
        wb_preview = load_workbook(tmp_preview_path, data_only=False, keep_links=True)
        ws_preview = wb_preview.active
        
        col_info = detect_column_info(ws_preview)
        
        col1, col2 = st.columns(2)
        
        with col1:
            product_model_options = [f"列{col_idx} - {get_column_name(ws_preview, col_idx, col_info.get('header_row_idx', 1))}" 
                                   for col_idx in range(1, ws_preview.max_column + 1)]
            
            default_product_model_idx = 0
            if col_info['product_model_col_idx']:
                default_product_model_idx = col_info['product_model_col_idx'] - 1
            
            product_model_column = st.selectbox(
                "产品型号列",
                options=product_model_options,
                index=default_product_model_idx,
                help="选择包含产品型号的列"
            )
        
        with col2:
            target_column_options = [f"列{col_idx} - {get_column_name(ws_preview, col_idx, col_info.get('header_row_idx', 1))}" 
                                    for col_idx in range(1, ws_preview.max_column + 1)]
            
            default_target_idx = 0
            if col_info['target_col_idx']:
                default_target_idx = col_info['target_col_idx'] - 1
            
            target_column_select = st.selectbox(
                "目标列（要填入数据的列）",
                options=target_column_options,
                index=default_target_idx,
                help="选择要更新数据的列"
            )
        
        data_start_row = st.number_input(
            "数据起始行",
            min_value=1,
            max_value=ws_preview.max_row,
            value=col_info.get('data_start_row', 4),
            help="数据行开始的行号（表头之后的第一个数据行）"
        )
        
        st.info(f"📊 表格信息: 共 {ws_preview.max_row} 行, {ws_preview.max_column} 列")
        
        st.session_state['preview_file_path'] = tmp_preview_path
        st.session_state['dist_file_ext'] = dist_file_ext
        
    except Exception as e:
        st.warning(f"⚠️ 无法预览文件: {str(e)}")
        st.warning("💡 请先点击'开始处理'按钮，系统会尝试自动识别列")
        os.unlink(tmp_preview_path)
else:
    st.info("💡 请先上传订单表，系统会自动识别列信息")

st.markdown("---")

st.markdown("### 🚀 开始处理")

if st.button("开始处理", type="primary", use_container_width=True):
    if not from_file:
        st.error("❌ 请先上传ERP库存表（from文件）")
        st.stop()
    
    if not dist_file:
        st.error("❌ 请先上传订单表（dist文件）")
        st.stop()
    
    if 'product_model_column' not in locals() or 'target_column_select' not in locals() or 'data_start_row' not in locals():
        st.error("❌ 请先上传订单表以配置列信息")
        st.stop()
    
    product_model_col_idx = int(product_model_column.split('-')[0].replace('列', ''))
    target_col_idx = int(target_column_select.split('-')[0].replace('列', ''))
    
    with st.spinner("正在处理数据..."):
        try:
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            status_text.text("📖 读取ERP库存表...")
            progress_bar.progress(10)
            
            try:
                df_source = pd.read_excel(from_file, header=1)
                status_text.text(f"✅ 成功读取ERP库存表，共 {len(df_source)} 行数据")
            except Exception as e:
                st.error(f"❌ 读取ERP库存表失败: {str(e)}")
                st.stop()
            
            progress_bar.progress(30)
            
            status_text.text("🔍 提取产品型号...")
            
            merchant_code_cols = [col for col in df_source.columns if '商家' in col and '编码' in col]
            
            if not merchant_code_cols:
                st.error("❌ 未在ERP库存表中找到商家编码列")
                st.stop()
            
            def extract_model(code):
                if isinstance(code, str) and '-' in code:
                    parts = code.split('-')
                    if len(parts) >= 2:
                        return '-'.join(parts[1:])
                return None
            
            for col in merchant_code_cols:
                df_source[f'产品型号_{col}'] = df_source[col].apply(extract_model)
            
            model_cols = [col for col in df_source.columns if '产品型号_' in col]
            
            if not model_cols:
                st.error("❌ 未能提取任何产品型号")
                st.stop()
            
            df_source['产品型号'] = df_source[model_cols[0]]
            for col in model_cols[1:]:
                df_source['产品型号'] = df_source['产品型号'].fillna(df_source[col])
            
            erp_models = set(df_source['产品型号'].dropna().unique())
            status_text.text(f"✅ 成功提取产品型号，共 {len(erp_models)} 个")
            progress_bar.progress(50)
            
            status_text.text("📊 计算差值...")
            
            if '实际可用数' not in df_source.columns or '30天销量' not in df_source.columns:
                st.error("❌ ERP库存表中缺少'实际可用数'或'30天销量'列")
                st.stop()
            
            df_source['差值'] = df_source['30天销量'] - df_source['实际可用数']
            status_text.text(f"✅ 成功计算差值")
            progress_bar.progress(60)
            
            status_text.text("📖 读取订单表...")
            
            try:
                if 'preview_file_path' in st.session_state and os.path.exists(st.session_state['preview_file_path']):
                    tmp_dist_path = st.session_state['preview_file_path']
                    file_ext = st.session_state.get('dist_file_ext', 'xlsx')
                else:
                    dist_content = dist_file.getvalue()
                    file_ext = dist_file.name.lower().split('.')[-1] if dist_file.name else 'xlsx'
                    
                    if file_ext == 'xls':
                        status_text.text("🔄 转换 .xls 为 .xlsx 格式...")
                        dist_content = convert_xls_to_xlsx_with_format(dist_content)
                    
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_dist:
                        tmp_dist.write(dist_content)
                        tmp_dist_path = tmp_dist.name
                
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_output:
                    tmp_output_path = tmp_output.name
                
                shutil.copy2(tmp_dist_path, tmp_output_path)
                
                wb = load_workbook(tmp_output_path, data_only=False, keep_links=True)
                ws = wb.active
                
                status_text.text(f"✅ 成功读取订单表，工作表名称: {ws.title}")
            except Exception as e:
                error_msg = str(e)
                if "does not support the old .xls file format" in error_msg.lower():
                    st.error("❌ 订单表文件格式不支持：")
                    st.error("openpyxl 库不支持旧的 .xls 文件格式")
                    st.info("💡 解决方案：请将 .xls 文件转换为 .xlsx 格式")
                    st.info("💡 转换方法：在 Excel 中打开文件，然后选择'文件 > 另存为 > Excel 工作簿 (.xlsx)'")
                elif "no valid workbook part" in error_msg.lower():
                    st.error("❌ 订单表文件格式不正确：")
                    st.error("该文件不是有效的 Excel (.xlsx) 格式")
                    st.info("💡 请确保上传的是 Excel 文件，而不是 CSV 或其他格式文件")
                    st.info("💡 如果是 CSV 文件，请先将其转换为 Excel 格式")
                else:
                    st.error(f"❌ 读取订单表失败: {error_msg}")
                st.stop()
            
            progress_bar.progress(70)
            
            status_text.text("🔍 使用配置的列信息...")
            
            st.info(f"📍 产品型号列: {product_model_column}")
            st.info(f"📍 目标列: {target_column_select}")
            st.info(f"📍 数据起始行: {data_start_row}")
            
            progress_bar.progress(80)
            
            status_text.text("🔄 更新数据...")
            
            model_diff_map = df_source.set_index('产品型号')['差值'].to_dict()
            
            st.info(f"📊 ERP库存表中产品型号数量: {len(model_diff_map)}")
            st.info(f"📊 ERP库存表中差值≥0的产品数量: {sum(1 for v in model_diff_map.values() if v >= 0)}")
            
            order_models = set()
            updated_count = 0
            skipped_count = 0
            matched_but_negative_count = 0
            
            for row in range(data_start_row, ws.max_row + 1):
                model = ws.cell(row=row, column=product_model_col_idx).value
                
                if model:
                    order_models.add(model)
                    if model in model_diff_map:
                        diff_value = model_diff_map[model]
                        if diff_value >= 0:
                            ws.cell(row=row, column=target_col_idx).value = diff_value
                            updated_count += 1
                        else:
                            matched_but_negative_count += 1
            
            st.info(f"📊 订单表中产品型号数量: {len(order_models)}")
            st.info(f"📊 匹配到但差值为负数的产品数量: {matched_but_negative_count}")
            status_text.text(f"✅ 数据更新完成，共更新了 {updated_count} 个单元格，跳过 {matched_but_negative_count} 个负数")
            progress_bar.progress(90)
            
            status_text.text("💾 保存文件...")
            wb.save(tmp_output_path)
            
            with open(tmp_output_path, 'rb') as f:
                st.session_state['output_file'] = f.read()
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            st.session_state['output_filename'] = f"订单表_更新_{timestamp}.xlsx"
            
            progress_bar.progress(100)
            status_text.text("✅ 处理完成！")
            
            st.success(f"🎉 处理成功！共更新了 {updated_count} 个产品型号，跳过 {skipped_count} 个负数")
            
            os.unlink(tmp_output_path)
            
            models_in_erp_not_in_order = sorted(erp_models - order_models)
            
            if models_in_erp_not_in_order:
                st.markdown("---")
                st.markdown("### ⚠️ ERP库存表中有但订单表中没有的产品型号")
                st.info(f"共找到 {len(models_in_erp_not_in_order)} 个产品型号在ERP库存表中存在，但在订单表中不存在：")
                
                def find_similar_model(target_model, all_models, threshold=0.8):
                    best_match = None
                    best_ratio = 0
                    for model in all_models:
                        ratio = difflib.SequenceMatcher(None, target_model, model).ratio()
                        if ratio >= threshold and ratio > best_ratio:
                            best_ratio = ratio
                            best_match = model
                    return best_match, best_ratio
                
                cols_per_row = 5
                for i in range(0, len(models_in_erp_not_in_order), cols_per_row):
                    cols = st.columns(cols_per_row)
                    for j, col in enumerate(cols):
                        if i + j < len(models_in_erp_not_in_order):
                            missing_model = models_in_erp_not_in_order[i + j]
                            similar_model, similarity = find_similar_model(missing_model, order_models)
                            
                            if similar_model:
                                col.markdown(f"**{missing_model}** → {similar_model} ({similarity*100:.0f}%)")
                            else:
                                col.markdown(f"**{missing_model}**")
            
        except Exception as e:
            st.error(f"❌ 处理过程中发生错误: {str(e)}")
            st.exception(e)
            st.stop()

st.markdown("---")

st.markdown("### 📥 下载结果")

if 'output_file' in st.session_state:
    st.download_button(
        label="📥 下载处理后的Excel文件",
        data=st.session_state['output_file'],
        file_name=st.session_state['output_filename'],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True
    )
    st.info(f"📄 文件名: {st.session_state['output_filename']}")
else:
    st.info("💡 请先上传文件并点击'开始处理'按钮")

st.markdown("---")

st.markdown("### 📋 使用说明")
st.markdown("""
1. **上传ERP库存表**：上传包含库存数据的Excel文件（支持.xlsx, .xls, .csv格式）
2. **上传订单表**：上传需要更新的订单Excel文件（支持.xlsx和.xls格式）
3. **配置列信息**：系统会自动识别产品型号列和目标列，您也可以手动选择
4. **开始处理**：点击按钮开始处理数据
5. **下载结果**：处理完成后，点击下载按钮获取更新后的Excel文件

**注意事项：**
- ERP库存表需要包含"实际可用数"和"30天销量"列
- 订单表需要包含产品型号列和目标列（系统会智能识别）
- 系统会自动提取产品型号并计算差值（30天销量 - 实际可用数）
- 只有非负数的差值才会填入订单表，负数会被跳过
- 处理后的文件会保留原始格式和图片
- 会显示ERP库存表中有但订单表中没有的产品型号
- 对于缺失的型号，会显示订单表中相似度最高的型号（相似度≥80%）
- 支持多种订单表格式，自动识别产品型号列（如：产品型号、商品货号、货号等）
- 支持多种目标列（如：所需数量、数量、订货数量、进货数量等）
- .xls格式文件会自动转换为.xlsx格式进行处理
""")